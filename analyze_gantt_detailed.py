import pdfplumber
import openpyxl
import os
import json

base_path = r"d:\Purdue\Courses\03. Spring 60300 - Graduate Planning\Other reference gantt chart"
output_file = r"d:\Purdue\Courses\03. Spring 60300 - Graduate Planning\gantt_analysis_output.txt"

output = []

def write(text=""):
    output.append(text)
    print(text)

# Analyze PDF files
pdf_files = [
    "Tara Middlebrooks_Gantt_PhD Timeline.pdf",
    "Trisolicha_GanttChart_TECH603_REV1.pdf",
    "_Gantt Chart.pdf"
]

write("="*100)
write("COMPREHENSIVE GANTT CHART FILE ANALYSIS")
write("="*100)

for pdf_file in pdf_files:
    pdf_path = os.path.join(base_path, pdf_file)
    if os.path.exists(pdf_path):
        write(f"\n{'='*100}")
        write(f"FILE: {pdf_file}")
        write(f"{'='*100}")
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                write(f"Pages: {len(pdf.pages)}")
                write(f"Size: {os.path.getsize(pdf_path) / 1024:.1f} KB")
                write(f"Creator: {pdf.metadata.get('Creator', 'Unknown') if pdf.metadata else 'N/A'}")
                
                for page_num, page in enumerate(pdf.pages, 1):
                    write(f"\n--- PAGE {page_num} ---")
                    write(f"Page dimensions: {page.width:.0f} x {page.height:.0f} points")
                    
                    # Text content
                    text = page.extract_text()
                    if text:
                        lines = [line.strip() for line in text.split('\n') if line.strip()]
                        write(f"\nText content ({len(lines)} lines):")
                        for idx, line in enumerate(lines, 1):
                            write(f"  {idx:2d}. {line}")
                    else:
                        write("No text content (image-based PDF)")
                    
                    # Tables
                    tables = page.extract_tables()
                    if tables:
                        write(f"\nTables found: {len(tables)}")
                        for ti, table in enumerate(tables, 1):
                            write(f"\n  TABLE {ti} ({len(table)} rows x {len(table[0]) if table else 0} cols):")
                            for row_idx, row in enumerate(table):
                                write(f"    Row {row_idx}: {row}")
                    
        except Exception as e:
            write(f"ERROR reading {pdf_file}: {str(e)}")

write(f"\n\n{'='*100}")
write("EXCEL FILE ANALYSIS")
write(f"{'='*100}")

excel_file = "TECH 603 Gantt chart.xlsx"
excel_path = os.path.join(base_path, excel_file)

if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        write(f"\nFile: {excel_file}")
        write(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            write(f"\n--- SHEET: '{sheet_name}' ---")
            write(f"Dimensions: {ws.dimensions}")
            write(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
            
            write(f"\nSheet content (all non-empty cells):")
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_info = f"[R{row_idx}:C{col_idx}]"
                        cell_value = str(cell.value)[:100]
                        cell_format = ""
                        if cell.fill and cell.fill.start_color:
                            cell_format += f" | Fill: {cell.fill.start_color.rgb}"
                        write(f"  {cell_info}: {cell_value}{cell_format}")
                        
        wb.close()
    except Exception as e:
        write(f"ERROR reading {excel_file}: {str(e)}")

write(f"\n{'='*100}")
write("ANALYSIS COMPLETE")
write(f"{'='*100}\n")

# Write to file
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print(f"\n\nAnalysis saved to: {output_file}")

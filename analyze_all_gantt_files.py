import pdfplumber
import openpyxl
import os
import json

base_path = r"d:\Purdue\Courses\03. Spring 60300 - Graduate Planning\Other reference gantt chart"

# Dictionary to store all analysis results
analysis_results = {}

# ============= ANALYZE PDF FILES =============
pdf_files = [
    "Tara Middlebrooks_Gantt_PhD Timeline.pdf",
    "Trisolicha_GanttChart_TECH603_REV1.pdf",
    "_Gantt Chart.pdf"
]

for pdf_file in pdf_files:
    pdf_path = os.path.join(base_path, pdf_file)
    if not os.path.exists(pdf_path):
        print(f"File not found: {pdf_file}")
        continue
        
    print(f"\n{'='*80}\nProcessing: {pdf_file}\n{'='*80}")
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            
            pdf_info = {
                'filename': pdf_file,
                'pages': len(pdf.pages),
                'size_kb': os.path.getsize(pdf_path) / 1024,
                'creator': pdf.metadata.get('Creator') if pdf.metadata else 'Unknown',
                'pages_data': []
            }
            
            for page_num, page in enumerate(pdf.pages, 1):
                print(f"Analyzing page {page_num}...")
                
                page_data = {
                    'page_num': page_num,
                    'width': page.width,
                    'height': page.height,
                    'text_lines': [],
                    'tables': [],
                    'shapes': len(page.lines) + len(page.rects) + len(page.curves)
                }
                
                # Extract text
                text = page.extract_text()
                if text:
                    lines = [l.strip() for l in text.split('\n') if l.strip()]
                    page_data['text_lines'] = lines
                    page_data['total_text_lines'] = len(lines)
                
                # Extract tables
                tables = page.extract_tables()
                if tables:
                    for table_idx, table in enumerate(tables, 1):
                        table_info = {
                            'table_num': table_idx,
                            'rows': len(table),
                            'cols': len(table[0]) if table else 0,
                            'content': table[:5]  # Store first 5 rows
                        }
                        page_data['tables'].append(table_info)
                
                pdf_info['pages_data'].append(page_data)
            
            analysis_results[pdf_file] = pdf_info
            print(f"Successfully processed {pdf_file}: {len(pdf.pages)} pages")
            
    except Exception as e:
        print(f"Error processing {pdf_file}: {str(e)}")

# ============= ANALYZE EXCEL FILE =============
excel_file = "TECH 603 Gantt chart.xlsx"
excel_path = os.path.join(base_path, excel_file)

print(f"\n{'='*80}\nProcessing: {excel_file}\n{'='*80}")

if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        excel_info = {
            'filename': excel_file,
            'sheets': wb.sheetnames,
            'sheets_data': {}
        }
        
        for sheet_name in wb.sheetnames:
            print(f"Analyzing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            sheet_data = {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_col': ws.max_column,
                'dimensions': str(ws.dimensions),
                'content': []
            }
            
            # Extract all content
            for row_idx in range(1, min(ws.max_row + 1, 101)):
                row_content = []
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        row_content.append({
                            'col': col_idx,
                            'value': str(cell.value)[:100],
                            'type': type(cell.value).__name__
                        })
                if row_content:
                    sheet_data['content'].append({
                        'row': row_idx,
                        'cells': row_content
                    })
            
            excel_info['sheets_data'][sheet_name] = sheet_data
        
        analysis_results[excel_file] = excel_info
        wb.close()
        print(f"Successfully processed {excel_file}: {len(wb.sheetnames)} sheets")
        
    except Exception as e:
        print(f"Error processing {excel_file}: {str(e)}")

# ============= OUTPUT RESULTS =============
print(f"\n{'='*80}\nOUTPUT SUMMARY\n{'='*80}\n")

for filename, data in analysis_results.items():
    print(f"\n{'='*80}")
    print(f"FILE: {filename}")
    print(f"{'='*80}")
    
    if 'pages' in data:  # PDF file
        print(f"Pages: {data['pages']}")
        print(f"Size: {data['size_kb']:.1f} KB")
        print(f"Creator: {data['creator']}")
        print(f"\nPage details:")
        for page_info in data['pages_data']:
            print(f"\n  Page {page_info['page_num']}:")
            print(f"    Dimensions: {page_info['width']:.0f} x {page_info['height']:.0f} pt")
            print(f"    Text lines: {page_info.get('total_text_lines', 0)}")
            print(f"    Tables: {len(page_info['tables'])}")
            print(f"    Shapes/graphical elements: {page_info['shapes']}")
            
            if page_info['text_lines']:
                print(f"    First 20 text lines:")
                for line in page_info['text_lines'][:20]:
                    print(f"      - {line[:70]}")
            
            if page_info['tables']:
                for table_info in page_info['tables'][:1]:
                    print(f"    First table ({table_info['rows']} rows x {table_info['cols']} cols):")
                    for row in table_info['content'][:3]:
                        print(f"      {row}")
    
    elif 'sheets' in data:  # Excel file
        print(f"Sheets: {', '.join(data['sheets'])}")
        
        for sheet_name, sheet_data in data['sheets_data'].items():
            print(f"\n  Sheet: {sheet_name}")
            print(f"    Dimensions: {sheet_data['dimensions']} ({sheet_data['max_row']} rows x {sheet_data['max_col']} cols)")
            print(f"    Non-empty cells: {len(sheet_data['content'])}")
            
            if sheet_data['content']:
                print(f"    First 10 rows of content:")
                for row_data in sheet_data['content'][:10]:
                    print(f"      Row {row_data['row']}: {[c['value'][:30] for c in row_data['cells']]}")

print(f"\n{'='*80}\nANALYSIS COMPLETE\n{'='*80}\n")

# Save to JSON file
json_output = r"d:\Purdue\Courses\03. Spring 60300 - Graduate Planning\gantt_analysis_results.json"
with open(json_output, 'w', encoding='utf-8') as f:
    # Convert non-serializable objects
    serializable_results = {}
    for key, val in analysis_results.items():
        serializable_results[key] = json.loads(json.dumps(val, default=str))
    json.dump(serializable_results, f, indent=2)

print(f"Detailed results saved to: {json_output}")

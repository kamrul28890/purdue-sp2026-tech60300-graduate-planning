import pdfplumber
import openpyxl
import os
from pathlib import Path
from pdf2image import convert_from_path
import json

base_path = r"d:\Purdue\Courses\03. Spring 60300 - Graduate Planning\Other reference gantt chart"

# Analyze PDF files
pdf_files = [
    "Tara Middlebrooks_Gantt_PhD Timeline.pdf",
    "Trisolicha_GanttChart_TECH603_REV1.pdf",
    "_Gantt Chart.pdf"
]

print("="*80)
print("PDF FILE ANALYSIS")
print("="*80)

for pdf_file in pdf_files:
    pdf_path = os.path.join(base_path, pdf_file)
    if os.path.exists(pdf_path):
        print(f"\n{'='*80}")
        print(f"FILE: {pdf_file}")
        print(f"{'='*80}")
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                print(f"Number of pages: {len(pdf.pages)}")
                print(f"PDF size: {os.path.getsize(pdf_path) / 1024:.1f} KB")
                
                # Get basic info
                if pdf.metadata:
                    print(f"Metadata keys: {list(pdf.metadata.keys())}")
                
                # Analyze each page
                for page_num, page in enumerate(pdf.pages, 1):
                    print(f"\n--- PAGE {page_num} ---")
                    print(f"Page size: {page.width} x {page.height} pt")
                    
                    # Extract text
                    text = page.extract_text()
                    if text:
                        text_lines = [line.strip() for line in text.split('\n') if line.strip()]
                        print(f"Total text lines: {len(text_lines)}")
                        print(f"First 25 lines of content:")
                        for line in text_lines[:25]:
                            print(f"  {line[:100]}")
                    else:
                        print("No text extracted (likely image-based PDF)")
                    
                    # Extract tables
                    tables = page.extract_tables()
                    if tables:
                        print(f"\nNumber of tables on page: {len(tables)}")
                        for ti, table in enumerate(tables, 1):
                            print(f"\n>>> TABLE {ti} <<<")
                            print(f"Dimensions: {len(table)} rows x {len(table[0]) if table else 0} cols")
                            for row_idx, row in enumerate(table[:10]):  # First 10 rows
                                print(f"  Row {row_idx}: {row}")
                    
                    # Get shapes/images
                    if hasattr(page, 'images') and page.images:
                        print(f"\nImages on page: {len(page.images)}")
                    
        except Exception as e:
            print(f"Error reading {pdf_file}: {str(e)}")

print("\n" + "="*80)
print("EXCEL FILE ANALYSIS")
print("="*80)

excel_file = "TECH 603 Gantt chart.xlsx"
excel_path = os.path.join(base_path, excel_file)

if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"\nFile: {excel_file}")
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- SHEET: '{sheet_name}' ---")
            print(f"Dimensions: {ws.dimensions}")
            print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
            
            # Get all content
            print(f"\nFull sheet content (non-empty cells):")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True), 1):
                non_empty = [str(cell)[:30] if cell is not None else "" for cell in row]
                if any(non_empty):
                    print(f"  Row {row_idx}: {non_empty}")
                    
        wb.close()
    except Exception as e:
        print(f"Error reading {excel_file}: {str(e)}")

print("\n" + "="*80)
print("ANALYSIS COMPLETE")
print("="*80)
